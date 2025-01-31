from datetime import datetime, timedelta
import os,random, pytz
from flask import Blueprint, current_app, jsonify, request
from flask_login import current_user, login_required
from openpyxl import load_workbook
from werkzeug.utils import secure_filename
from models import Attendance, Class, Student, Teacher, db
from helpers.validation import allowed_file, validate_student_data
from helpers.response import handle_error, make_response

api_bp = Blueprint('api', __name__)

@api_bp.route("/api/add-students-bulk", methods=["POST"])
@login_required
def add_students_bulk():
    class_name = request.form['class_name']
    if 'file' not in request.files:
        return handle_error("No file part in the request", 400)
    file = request.files['file']
    if file.filename == '':
        return handle_error("No selected file", 400)
    if file and allowed_file(file.filename):
        filename = secure_filename(file.filename)
        upload_folder = current_app.config['UPLOAD_FOLDER']
        filepath = os.path.join(upload_folder, filename)
        if not os.path.exists(upload_folder):
            os.makedirs(upload_folder)
        file.save(filepath)
        
        try:
            wb = load_workbook(filepath)
            sheet = wb.active
            students_added = []
            for row in sheet.iter_rows(min_row=2, values_only=True):
                sr_no, roll_no, name = row
                if roll_no and name:
                    class_instance = Class.query.filter_by(name=class_name).first()
                    if not class_instance:
                        return handle_error(f"Class '{class_name}' not found.", 400)
                    new_student = Student(name=name, roll_no=roll_no, class_id=class_instance.id)
                    db.session.add(new_student)
                    students_added.append({"roll_no": roll_no, "name": name})

            db.session.commit()
            return make_response(True, "Students added successfully", {"students": students_added}, 200)
        except Exception as e:
            db.session.rollback()
            return handle_error(str(e), 500)
        finally:
            os.remove(filepath)
    else:
        return handle_error("Invalid file format", 400)

@api_bp.route("/api/add-student", methods=["POST"])
@login_required
def add_student():
    try:
        data = request.get_json()
        errors = validate_student_data(data)
        if errors:
            return handle_error(errors, 400)
        full_name = data["full_name"]
        class_name = data["class_name"]
        roll_no = data["roll_no"]
        class_instance = Class.query.filter_by(name=class_name).first()
        if not class_instance:
            return handle_error(f"Class '{class_name}' not found.", 400)
        new_student = Student(name=full_name, roll_no=roll_no, class_id=class_instance.id)
        db.session.add(new_student)
        db.session.commit()
        return make_response(True, f"Student '{full_name}' added to class '{class_name}'")
    except Exception as e:
        return handle_error(str(e), 500)

@api_bp.route("/api/edit-student/<int:student_id>", methods=["POST"])
@login_required
def edit_student(student_id):
    data = request.get_json()
    student = Student.query.get(student_id)
    if student:
        student.name = data["name"]
        student.roll_no = data["roll_no"]
        class_name = data["class_name"]
        student_class = Class.query.filter_by(name=class_name).first()
        if student_class:
            student.class_id = student_class.id
        else:
            return handle_error("Class not found.", 404)
        db.session.commit()
        return make_response(True, "Student updated successfully.", {"id": student.id, "name": student.name, "roll_no": student.roll_no, "class_name": class_name}, 200)
    return handle_error("Student not found.", 404)

@api_bp.route("/api/delete-student/<int:student_id>", methods=["DELETE"])
@login_required
def delete_student(student_id):
    student = Student.query.get(student_id)
    if student:
        db.session.delete(student)
        db.session.commit()
        return make_response(True, "Student deleted successfully.", None, 200)
    return handle_error("Student not found.", 404)

@api_bp.route("/api/students")
@login_required
def search_students():
    class_name = request.args.get("class_name")
    query = Student.query
    if class_name:
        student_class = Class.query.filter_by(name=class_name).first()
        if not student_class:
            return handle_error("Class not found.", 404)
        query = query.filter_by(class_id=student_class.id)
    students = query.all()
    return make_response(True, "Students retrieved successfully.", {"students": [{"id": student.id, "name": student.name, "class_name": class_name, "roll_no": student.roll_no} for student in students]}, 200)

@api_bp.route("/attendance", methods=["GET"])
def get_attendance():
    date_str = request.args.get("date")
    class_name = request.args.get("class_name")
    subject = request.args.get("subject")
    if not date_str or not class_name or not subject:
        return handle_error("Date, subject & class are required", 400)
    date = datetime.strptime(date_str, "%Y-%m-%d").date()
    student_class = Class.query.filter_by(name=class_name).first()
    if not student_class:
        return handle_error("Class not found", 404)
    attendance_records = Attendance.query.join(Student).filter(Attendance.date == date, Student.class_id == student_class.id, Attendance.subject == subject).all()
    if len(attendance_records) == 0:
        return handle_error("No records were found.", 404)
    records = [{"id": record.student.id, "roll_no": record.student.roll_no, "name": record.student.name, "present": record.present} for record in attendance_records]
    return make_response(True, "Attendance records retrieved successfully.", records, 200)

@api_bp.route("/api/attendance/students", methods=["GET"])
@login_required
def get_students_by_grade():
    class_name = request.args.get("class_name")
    class_instance = Class.query.filter_by(name=class_name).first()
    if not class_instance:
        return handle_error(f"Class '{class_name}' not found.", 400)
    students = Student.query.filter_by(class_id=class_instance.id).all()
    if not students:
        return handle_error("No students found for the given class.", 404)
    current_date = datetime.now(pytz.timezone('Asia/Kolkata')).date()
    students_data = []
    for student in students:
        attendance_record = next((att for att in student.attendance_records if att.date == current_date), None)
        present_status = attendance_record.present if attendance_record else False
        students_data.append({"id": student.id, "roll_no": student.roll_no, "name": student.name, "present": present_status})
    return make_response(True, "Students retrieved successfully.", students_data, 200)
@api_bp.route('/api/attendance', methods=['POST'])
@login_required
def update_attendance():
    data = request.json
    class_name = data.get('class_name')
    attendance_data = data.get('attendance')
    subject = data.get('subject')
    if subject not in current_user.subjects:
        return handle_error("You are not authorized to handle this subject.", 403)
    class_instance = Class.query.filter_by(name=class_name).first()
    if not class_instance:
        return handle_error(f"Class '{class_name}' not found.", 400)
    assigned_class = Class.query.join(Teacher).filter(Class.id == class_instance.id, Teacher.id == current_user.id).first()
    if not assigned_class:
        return handle_error("You do not have permission to modify/view attendance for this class.", 403)
    for record in attendance_data:
        student_id = record.get('student_id')
        present = record.get('present')
        note = record.get('note', f'{current_user.name} - {assigned_class.name}')
        student = Student.query.filter_by(id=student_id, class_id=class_instance.id).first()
        if not student: continue
        existing_record = Attendance.query.filter_by(student_id=student.id, date=datetime.now(pytz.timezone('Asia/Kolkata')).date(), teacher_id=current_user.id, note=f'{current_user.name} - {assigned_class.name}', subject=subject).first()
        if existing_record is None:
            attendance = Attendance(student_id=student.id, date=datetime.now(pytz.timezone('Asia/Kolkata')).date(), present=present, note=note, teacher_id=current_user.id, subject=subject)
            db.session.add(attendance)
    db.session.commit()
    return make_response(True, "Attendance updated successfully", None, 200)

@api_bp.route("/attendance/low-attendance", methods=["GET"])
@login_required
def get_low_attendance_students():
    class_name = request.args.get("class_name")
    subject = request.args.get("subject")
    year = request.args.get("year", datetime.now().year)
    month = request.args.get("month", datetime.now().month)
    if not class_name or not subject:
        return handle_error("Class name and subject are required", 400)
    if subject not in current_user.subjects:
        return handle_error("You are not authorized to handle this subject.", 403)
    class_instance = Class.query.filter_by(name=class_name).first()
    if not class_instance:
        return handle_error(f"Class '{class_name}' not found.", 400)
    assigned_class = Class.query.join(Teacher).filter(Class.id == class_instance.id, Teacher.id == current_user.id).first()
    if not assigned_class:
        return handle_error("You do not have permission to modify/view attendance for this class.", 403)
    students = Student.query.filter_by(class_id=class_instance.id).all()
    low_attendance_students = []
    for student in students:
        total_attendance_days = Attendance.query.filter(Attendance.student_id == student.id, Attendance.subject == subject, db.extract('year', Attendance.date) == year, db.extract('month', Attendance.date) == month).count()
        present_days = Attendance.query.filter(Attendance.student_id == student.id, Attendance.present == True, Attendance.subject == subject, db.extract('year', Attendance.date) == year, db.extract('month', Attendance.date) == month).count()
        if total_attendance_days > 0:
            attendance_percentage = (present_days / total_attendance_days) * 100
            if attendance_percentage < 50:
                low_attendance_students.append({"id": student.id, "roll_no": student.roll_no, "name": student.name, "attendance_percentage": attendance_percentage})
    return make_response(True, "Low attendance students retrieved successfully.", {"students": low_attendance_students}, 200)

@api_bp.route("/attendance/populate-dummy", methods=["POST"])
@login_required
def populate_dummy_data():
    class_name = request.form.get("class_name")
    subject = request.form.get("subject")
    teacher_id = current_user.id
    days_in_month = int(request.form.get("days", 30))
    if not class_name or not subject:
        return jsonify({"error": "Class name and subject are required"}), 400
    class_instance = Class.query.filter_by(name=class_name).first()
    assigned_class = Class.query.join(Teacher).filter(Class.id == class_instance.id, Teacher.id == current_user.id).first()
    if not assigned_class:
        abort(403, description="You do not have permission to modify/view attendance for this class.")
    if not class_instance:
        return jsonify({"success": False, "message": f"Class '{class_name}' not found."}), 400
    students = Student.query.filter_by(class_id=class_instance.id).all()
    try:
        for day_offset in range(days_in_month):
            date = datetime.now() - timedelta(days=(days_in_month - day_offset))
            for student in students:
                present = random.choice([True, False])
                new_attendance = Attendance(date=date.date(), present=present, subject=subject, student_id=student.id, teacher_id=teacher_id)
                db.session.add(new_attendance)
        db.session.commit()
        return jsonify({"success": True, "message": "Dummy data added successfully"}), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({"success": False, "message": str(e)}), 500
